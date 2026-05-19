```python
import re
from datetime import datetime
from urllib.parse import urlparse, parse_qs

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pyscript import document
from js import Blob, URL
import base64
from io import BytesIO

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

output = document.getElementById("output")

def log(text):
    output.innerText += str(text) + "\n"

def extract_scholar_id(input_str: str) -> str:
    input_str = input_str.strip()

    if input_str.startswith("http"):
        parsed = urlparse(input_str)
        qs = parse_qs(parsed.query)

        if "user" in qs:
            return qs["user"][0]

        raise ValueError("URL tidak mengandung parameter user")

    return input_str

def fetch_scholar_data(scholar_id: str):

    base_url = "https://scholar.google.com/citations"

    params = {
        "user": scholar_id,
        "sortby": "pubdate",
        "cstart": 0,
        "pagesize": 100,
    }

    log(f"Mengambil data Scholar ID: {scholar_id}")

    resp = requests.get(
        base_url,
        params=params,
        headers=HEADERS
    )

    soup = BeautifulSoup(resp.text, "html.parser")

    name_tag = soup.find("div", id="gsc_prf_in")
    author_name = name_tag.get_text(strip=True) if name_tag else "Unknown"

    aff_tag = soup.find("div", class_="gsc_prf_il")
    affiliation = aff_tag.get_text(strip=True) if aff_tag else ""

    stats_cells = soup.select("#gsc_rsb_st td.gsc_rsb_std")

    total_citations = stats_cells[0].get_text(strip=True) if len(stats_cells) > 0 else "0"
    h_index = stats_cells[2].get_text(strip=True) if len(stats_cells) > 2 else "0"
    i10_index = stats_cells[4].get_text(strip=True) if len(stats_cells) > 4 else "0"

    citations_per_year = {}

    years = soup.select(".gsc_g_t")
    scores = soup.select(".gsc_g_al")

    for y, s in zip(years, scores):
        citations_per_year[y.get_text(strip=True)] = s.get_text(strip=True)

    publications = []

    for row in soup.select("#gsc_a_t .gsc_a_tr"):

        title_tag = row.select_one(".gsc_a_at")

        title = title_tag.get_text(strip=True) if title_tag else ""

        link = (
            "https://scholar.google.com" + title_tag["href"]
            if title_tag and title_tag.get("href")
            else ""
        )

        gray = row.select(".gs_gray")

        authors = gray[0].get_text(strip=True) if len(gray) > 0 else ""
        venue = gray[1].get_text(strip=True) if len(gray) > 1 else ""

        cite_tag = row.select_one(".gsc_a_ac")
        cite_text = cite_tag.get_text(strip=True) if cite_tag else ""

        citations = int(cite_text) if cite_text.isdigit() else 0

        year_tag = row.select_one(".gsc_a_h")
        year_text = year_tag.get_text(strip=True) if year_tag else ""

        year = int(year_text) if year_text.isdigit() else 0

        publications.append({
            "title": title,
            "authors": authors,
            "venue": venue,
            "link": link,
            "citations": citations,
            "year": year,
        })

    return {
        "scholar_id": scholar_id,
        "author_name": author_name,
        "affiliation": affiliation,
        "total_citations": total_citations,
        "h_index": h_index,
        "i10_index": i10_index,
        "citations_per_year": citations_per_year,
        "publications": publications,
    }

def save_excel(data):

    wb = Workbook()

    ws = wb.active
    ws.title = "Profil"

    ws["A1"] = "PROFIL GOOGLE SCHOLAR"

    ws["A3"] = "Nama"
    ws["B3"] = data["author_name"]

    ws["A4"] = "Institusi"
    ws["B4"] = data["affiliation"]

    ws["A5"] = "Total Sitasi"
    ws["B5"] = data["total_citations"]

    ws2 = wb.create_sheet("Publikasi")

    headers = [
        "No",
        "Judul",
        "Penulis",
        "Venue",
        "Sitasi",
        "Tahun"
    ]

    for col, h in enumerate(headers, start=1):
        ws2.cell(1, col, h)

    for i, pub in enumerate(data["publications"], start=2):

        ws2.cell(i, 1, i-1)
        ws2.cell(i, 2, pub["title"])
        ws2.cell(i, 3, pub["authors"])
        ws2.cell(i, 4, pub["venue"])
        ws2.cell(i, 5, pub["citations"])
        ws2.cell(i, 6, pub["year"])

    bio = BytesIO()

    wb.save(bio)

    bio.seek(0)

    data64 = base64.b64encode(
        bio.read()
    ).decode()

    filename = (
        "scholar_" +
        re.sub(r"[^\w\-]", "_", data["author_name"]) +
        ".xlsx"
    )

    js_code = f"""
    const link = document.createElement('a');
    link.href = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{data64}";
    link.download = "{filename}";
    link.click();
    """

    from js import eval
    eval(js_code)

def run_scraper(event=None):

    try:

        output.innerText = ""

        input_el = document.getElementById("scholarInput")

        user_input = input_el.value.strip()

        scholar_id = extract_scholar_id(user_input)

        log(f"Scholar ID: {scholar_id}")

        data = fetch_scholar_data(scholar_id)

        log(f"Nama: {data['author_name']}")
        log(f"Sitasi: {data['total_citations']}")
        log(f"Publikasi: {len(data['publications'])}")

        save_excel(data)

        log("Excel berhasil dibuat")

    except Exception as e:

        log(f"ERROR: {str(e)}")
```
